from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from sj_generator.io.excel_repo import save_questions
from sj_generator.io.export_md import export_questions_to_markdown
from sj_generator.models import Question


@dataclass(frozen=True)
class BatchFolderizeResult:
    source_path: Path
    target_dir: Path
    target_xlsx: Path
    target_md: Path
    question_count: int


def process_excel_to_folder_mode(path: Path, *, export_date: date | None = None) -> BatchFolderizeResult:
    src = Path(path)
    if not src.exists():
        raise FileNotFoundError(src)
    if src.suffix.lower() != ".xlsx":
        raise ValueError(f"仅支持 xlsx 文件：{src}")

    target_dir = src.parent / src.stem
    target_dir.mkdir(parents=True, exist_ok=True)
    target_xlsx = target_dir / src.name
    if src.resolve() != target_xlsx.resolve():
        if target_xlsx.exists():
            target_xlsx.unlink()
        src.replace(target_xlsx)

    questions = load_compatible_questions(target_xlsx)
    save_questions(target_xlsx, questions)

    target_md = target_dir / f"{target_xlsx.stem}.md"
    md_text = export_questions_to_markdown(
        excel_file_name=target_xlsx.stem,
        export_date=export_date or date.today(),
        questions=questions,
    )
    target_md.write_text(md_text, encoding="utf-8")
    return BatchFolderizeResult(
        source_path=src,
        target_dir=target_dir,
        target_xlsx=target_xlsx,
        target_md=target_md,
        question_count=len(questions),
    )


def process_excel_files_to_folder_mode(
    paths: list[Path], *, export_date: date | None = None
) -> list[BatchFolderizeResult]:
    return [process_excel_to_folder_mode(path, export_date=export_date) for path in paths]


def load_compatible_questions(path: Path) -> list[Question]:
    wb = load_workbook(path)
    ws = wb["questions"] if "questions" in wb.sheetnames else wb.active
    header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_map: dict[str, int] = {}
    for idx, value in enumerate(header_row):
        if isinstance(value, str) and value.strip():
            header_map[value.strip()] = idx
    required = ["编号", "题目", "答案", "解析"]
    missing = [name for name in required if name not in header_map]
    if missing:
        raise ValueError(f"题库表头缺少列：{', '.join(missing)}")

    result: list[Question] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        number = _to_text(row[header_map["编号"]])
        stem = _to_text(row[header_map["题目"]])
        options = _to_text(row[header_map["选项"]]) if "选项" in header_map else ""
        answer = _to_text(row[header_map["答案"]])
        analysis = _to_text(row[header_map["解析"]])
        if not any([number, stem, options, answer, analysis]):
            continue
        result.append(
            Question(
                number=number,
                stem=stem,
                options=options,
                answer=answer,
                analysis=analysis,
            )
        )
    return result


def _to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()
