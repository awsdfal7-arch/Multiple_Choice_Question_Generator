from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from sj_generator.models import Question


SHEET_NAME = "questions"
HEADERS = ["编号", "题目", "选项", "答案", "解析"]


def create_empty_repo(path: Path) -> None:
    if path.exists():
        raise FileExistsError(f"文件已存在：{path}")
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(HEADERS)
    wb.save(path)


def load_questions(path: Path) -> list[Question]:
    wb = load_workbook(path)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

    header_row, header_map = _get_header_map(ws)
    _ensure_required_headers(header_map)

    result: list[Question] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        number = _cell_to_str(row[header_map["编号"]])
        stem = _cell_to_str(row[header_map["题目"]])
        options = _cell_to_str(row[header_map["选项"]])
        answer = _cell_to_str(row[header_map["答案"]])
        analysis = _cell_to_str(row[header_map["解析"]])
        if not any([number, stem, options, answer, analysis]):
            continue
        result.append(
            Question(
                number=number,
                stem=stem,
                options=options,
                answer=answer,
                analysis=analysis,
            )
        )
    return result


def append_questions(path: Path, questions: list[Question]) -> None:
    wb = load_workbook(path)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

    header_row, header_map = _get_header_map(ws)
    _ensure_required_headers(header_map)

    for q in questions:
        row: list[Any] = [None] * len(header_row)
        row[header_map["编号"]] = q.number
        row[header_map["题目"]] = q.stem
        row[header_map["选项"]] = q.options
        row[header_map["答案"]] = q.answer
        row[header_map["解析"]] = q.analysis
        ws.append(row)

    wb.save(path)

def save_questions(path: Path, questions: list[Question]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(HEADERS)
    for q in questions:
        ws.append([q.number, q.stem, q.options, q.answer, q.analysis])
    wb.save(path)


def try_parse_options_json(options: str) -> dict[str, str] | None:
    raw = options.strip()
    if not raw.startswith("{") or not raw.endswith("}"):
        return None
    try:
        obj = json.loads(raw)
    except Exception:
        return None
    if not isinstance(obj, dict):
        return None
    normalized: dict[str, str] = {}
    for k, v in obj.items():
        if not isinstance(k, str):
            return None
        if not isinstance(v, str):
            v = str(v)
        normalized[k.strip()] = v
    if not normalized:
        return None
    return normalized


def validate_repo(path: Path) -> None:
    wb = load_workbook(path)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    _, header_map = _get_header_map(ws)
    _ensure_required_headers(header_map)


def _ensure_required_headers(header_map: dict[str, int]) -> None:
    missing = [h for h in HEADERS if h not in header_map]
    if missing:
        raise ValueError(f"题库表头缺少列：{', '.join(missing)}")


def _get_header_map(ws: Any) -> tuple[list[Any], dict[str, int]]:
    header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_map: dict[str, int] = {}
    for idx, v in enumerate(header_row):
        if isinstance(v, str) and v.strip():
            header_map[v.strip()] = idx
    return header_row, header_map


def _cell_to_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    return str(v).strip()
